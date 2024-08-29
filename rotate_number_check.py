from paddleocr import PaddleOCR
from PIL import Image, ImageDraw, ImageEnhance, ImageFilter
import cv2
import numpy as np
import glob
import os
import openpyxl
import pytesseract
import logging

logging.getLogger('ppocr').setLevel(logging.ERROR)

# PaddleOCR 리더 생성 (한글 'korean' 설정)
ocr = PaddleOCR(lang='korean')

# 입력 폴더 경로 및 출력 폴더 설정
input_folder = './sample' # 원하는 폴더로 수정 필요
output_folder = f'{os.path.basename(os.path.normpath(input_folder))}_ans'
os.makedirs(output_folder, exist_ok=True)

image_paths = glob.glob(os.path.join(input_folder, '*.jpg'))  # jpg 파일만 가져오는 예시

def image_check(image, region):
    # 이미지 크기 가져오기
    width, height = image.size

    # 영역 설정 (우측 하단, 중간 하단, 좌측 하단)
    if region == 'right':
        region_x = int(width * 0.75)
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, width, height))
    elif region == 'center':
        region_x = int(width * 0.25)
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, region_x + int(width * 0.5), height))
    elif region == 'left':
        region_x = 0
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, region_x + int(width * 0.25), height))

    # 텍스트 검출
    cropped_image_np = np.array(cropped_image)
    result = ocr.ocr(cropped_image_np, cls=True)

    if result[0] is None:
        result[0] = [[[[1.0, 5.0], [38.0, 5.0], [36.0, 317.0], [0.0, 316.0]], ('숫자', 0.6)]]
    
    # PIL의 이미지를 수정하기 위한 객체 생성
    draw = ImageDraw.Draw(image)
    # 결과 시각화
    boxes = [line[0] for line in result[0]]
    texts = [line[1][0] for line in result[0]]
    scores = [line[1][1] for line in result[0]]

    text_num = False
    text_num2 = False
    
    # 이미지에 박스와 텍스트 그리기
    for (box, text, score) in zip(boxes, texts, scores):
        if score > 0.8:
            if text.isdigit():  # 인식된 텍스트가 숫자인지 확인
                text_num = True
                (top_left, top_right, bottom_right, bottom_left) = box

                # 좌표를 원본 이미지의 좌표로 변환
                top_left = (int(top_left[0] + region_x), int(top_left[1] + region_y))
                bottom_right = (int(bottom_right[0] + region_x), int(bottom_right[1] + region_y))
                bottom_left = (int(bottom_left[0] + region_x), int(bottom_left[1] + region_y))

                # 박스 그리기
                if int(text) == int(filename):
                    print(int(filename), '에서 파일명과 동일한 ', int(text), '를 찾았습니다.')
                    draw.rectangle([top_left, bottom_right], outline=(0, 255, 0), width=2)
                    draw.text(bottom_left, text, fill=(255, 0, 0))  # 빨간색 텍스트
                    text_num2 = True
    
    return text_num, text_num2


def image_check_single(image, region):
    width, height = image.size

    # 영역 설정 (우측 하단, 중간 하단, 좌측 하단)
    if region == 'right':
        region_x = int(width * 0.75)
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, width, height))
    elif region == 'center':
        region_x = int(width * 0.25)
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, region_x + int(width * 0.5), height))
    elif region == 'left':
        region_x = 0
        region_y = int(height * 0.75)
        cropped_image = image.crop((region_x, region_y, region_x + int(width * 0.25), height))

    # 이미지 전처리 강화
    gray_image = cropped_image.convert('L')
    enhancer = ImageEnhance.Contrast(gray_image)
    enhanced_image = enhancer.enhance(2)
    filtered_image = enhanced_image.filter(ImageFilter.MedianFilter())

    # OCR 수행 (여러 설정으로 시도)
    configs = [
        r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789',
        r'--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789',
        r'--oem 3 --psm 8 -c tessedit_char_whitelist=0123456789',
        r'--oem 3 --psm 13 -c tessedit_char_whitelist=0123456789'
    ]

    text_num = False
    text_num2 = False

    for config in configs:
        text = pytesseract.image_to_string(filtered_image, config=config)
        lines = text.splitlines()

        for line in lines:
            # 결과 정제
            cleaned_line = ''.join(filter(str.isdigit, line))
            if cleaned_line:
                try:
                    if int(cleaned_line) == int(filename):
                        print(f"{int(filename)}에서 파일명과 동일한 {cleaned_line}을 찾았습니다.")
                        text_num = True
                        text_num2 = True
                        return text_num, text_num2
                except ValueError:
                    pass

    return text_num, text_num2

well_files = []     # 올바른 파일 리스트를 저장할 리스트
rotate_files = []   # 회전이 일어난 파일 리스트를 저장할 리스트
rotate_angles = []  # 회전이 일어난 각도 리스트를 저장할 리스트
no_mark_files = []  # 면표시가 없는 파일 리스트를 저장할 리스트

regions = ['right', 'center', 'left']

for image_path in image_paths:
    image = Image.open(image_path)
    filename = os.path.splitext(os.path.basename(image_path))[0]
    filename_non_zero = filename.lstrip('0')

    text_detected = False  #z text_detected 변수를 루프 시작 시 초기화

    for region in regions:
        for angle in [0, 90, 180, 270]:  # 90도씩 회전하며 검사
            rotated_image = image.rotate(angle, expand=True)
            
            if len(filename_non_zero)<2:
                image_check_num, image_check_num2 = image_check_single(rotated_image, region)
            else:
                image_check_num, image_check_num2 = image_check(rotated_image, region)
            
            if image_check_num:
                output_path = os.path.join(output_folder, os.path.basename(image_path))
                rotated_image.save(output_path)  # 텍스트가 감지되면 이미지 저장
                text_detected = True
                
            if image_check_num2:
                if angle == 0:
                    well_files.append(image_path)
                else:
                    rotate_files.append(image_path)
                    rotate_angles.append(angle)
                break
        if image_check_num2:
            break
    
    if not text_detected or not image_check_num2:
        # 360도 회전 후에도 텍스트가 없으면 원본 상태로 복원
        image = image.rotate(0)  # 원래 상태로 되돌림
        output_path = os.path.join(output_folder, os.path.basename(image_path))
        image.save(output_path)
        no_mark_files.append(image_path)  # 면표시가 없는 파일을 리스트에 추가

# 새로운 엑셀 파일 생성
workbook = openpyxl.Workbook()

# 시트 생성 및 데이터 저장
# 올바르게 인식한 파일 리스트를 "Well Files" 시트에 저장
well_files_sheet = workbook.active
well_files_sheet.title = "인식 완료"
for index, well_file in enumerate(well_files, start=1):
    well_files_sheet.cell(row=index, column=1).value = well_file

# 회전이 일어난 파일 리스트를 "회전 후 인식 완료" 시트에 저장
rotate_files_sheet = workbook.create_sheet(title="회전 후 인식 완료")
for index, (rotate_file, rotate_angle) in enumerate(zip(rotate_files, rotate_angles), start=1):
    # 첫 번째 열에 파일 이름 저장
    rotate_files_sheet.cell(row=index, column=1).value = rotate_file
    # 세 번째 열에 회전 각도 저장
    rotate_files_sheet.cell(row=index, column=3).value = rotate_angle


# 면표시가 없는 파일 리스트를 "No Mark Files" 시트에 저장
no_mark_files_sheet = workbook.create_sheet(title="인식 실패")
for index, no_mark_file in enumerate(no_mark_files, start=1):
    no_mark_files_sheet.cell(row=index, column=1).value = no_mark_file

# 엑셀 파일 저장
workbook.save(f'{input_folder}.xlsx')
